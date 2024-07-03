import datetime
import os
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import simplejson
from flask import current_app
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame as df
from sqlalchemy import func, text, select, Integer, Float, String, case, and_
from sqlalchemy.orm import aliased, Session

from app.g_sheets import gSheets
from app.models import MessageOrders, Messages, Goods, Settings, Clients, Customers, Itog, Payers, Payments, \
    customers_to_clients, payers_to_clients, Prices
from datetime import datetime as dt
import json


class Reports:

    def __init__(self, start_date=None, end_date=None):
        self.org_half=100
        self.start_date = start_date
        self.end_date = end_date

    def get_itog(self):
        session: Session = current_app.session
        start_date = session.query(Settings).filter(Settings.name == Settings.START_DATE).one()
        start_date = dt.strptime(start_date.value, "%d.%m.%Y")

        payments_subquery = select(func.coalesce(Payments.for_client_id, Clients.id).label('p_client_id'),
                                   func.sum(Payments.sum).label('p_sum')).join(Payers).join(Clients, Payers.clients) \
            .filter(Payments.timestamp >= start_date, func.coalesce(Payments.not_use, False) == False).group_by(
            text("1")).subquery()

        query = select(Clients.name, func.sum(Itog.org), func.sum(Itog.sum), func.sum(payments_subquery.c.p_sum),
                       func.sum(Itog.sum - payments_subquery.c.p_sum)) \
            .join(Clients).join(payments_subquery, Itog.client_id == payments_subquery.c.p_client_id, isouter=True) \
            .filter(Itog.date == start_date, Itog.good_id == None).group_by(text("1"))
        result_q = session.execute(query)
        result = []
        for row in result_q.mappings():
            result.append(dict(row))
        return result

    def getMessages(self, client_id, good_id=None):
        if self.start_date is None:
            start_date = current_app.session.query(Settings).filter(Settings.name == Settings.START_DATE).one()
            start_date = dt.strptime(start_date.value, "%d.%m.%Y")
        else:
            start_date = self.start_date
        good_id = int(good_id) if good_id else None
        """получаем заказчика, чтобы знать кому слать сообщения"""
        customer = dict(next(current_app.session.execute(
            select(Customers.id, Customers.wa_id,Customers.name,Customers.number,Customers.push_name, Customers.params).join(Clients, Customers.clients).filter(Clients.id == client_id)
        ).mappings()))
        """подзапрос для группировки для общего кол-ва для получения всего заказа"""
        query = select(Goods.name, func.sum(MessageOrders.quantity).label("sum")) \
            .select_from(Messages).join(MessageOrders, isouter=True).join(Goods, isouter=True) \
            .join(Customers).join(customers_to_clients, isouter=True).join(Clients,
                                                                           func.coalesce(Messages.for_client_id,
                                                                                         customers_to_clients.c.client_id) == Clients.id) \
            .filter(Messages.timestamp >= start_date,
                    func.coalesce(Clients.duplicate_for, Clients.id) == int(client_id)).group_by(text("1"))
        if self.end_date:
            query = query.filter(Messages.timestamp < self.end_date)
        subq = query.subquery()
        '''запрос для всего заказа (суммируем все сообщения)'''
        order = next(current_app.session.execute(select(func.aggregate_strings(subq.c.name + '-' + func.cast(subq.c.sum, String), ';').label('order')).select_from(subq)).mappings()).order
        """выбираем сообщения для таблицы"""
        query = select(Messages.id, Messages.text.label("Сообщение"), Messages.props['comment'].as_string().label("Коммент"),
                       func.strftime('%d.%m.%Y %H:%M', Messages.timestamp).label("Дата"), func.aggregate_strings(Goods.name+'-'+func.cast(MessageOrders.quantity, String), ';').label("Заказ"))\
            .select_from(Messages).join(MessageOrders, isouter=True).join(Goods, isouter=True) \
            .join(Customers).join(customers_to_clients, isouter=True).join(Clients, func.coalesce(Messages.for_client_id, customers_to_clients.c.client_id) == Clients.id) \
            .filter(Messages.timestamp >= start_date, func.coalesce(Clients.duplicate_for, Clients.id) == int(client_id))
        if self.end_date:
            query = query.filter(Messages.timestamp < self.end_date)
        if good_id:
           query = query.filter(Goods.id == func.coalesce(good_id, Goods.id))
        query = query.order_by(Messages.timestamp)
        query = query.group_by(text("1, 2, 3"))
        # todo переделать
        case_qty = case(
            ((MessageOrders.quantity / 0.5 % 2) == 0,MessageOrders.quantity),
            ((MessageOrders.quantity / 0.5 % 2) != 0, MessageOrders.quantity-0.5),
        else_=0)
        # есть ли половинка
        case_half_qty = case(
            ((MessageOrders.quantity / 0.5 % 2) != 0, 1),
        else_=0)
        '''запрос общей суммы'''
        sum_query = select(
                      func.cast(func.sum(MessageOrders.quantity * Goods.price + case_qty * Goods.org_price + case_half_qty * 100), Float).label('sum')) \
            .select_from(MessageOrders) \
            .join(Messages).join(Customers).join(customers_to_clients, isouter=True).join(Goods) \
            .filter(Messages.timestamp >= start_date, func.coalesce(Messages.for_client_id, customers_to_clients.c.client_id) == int(client_id))
        if self.end_date:
            sum_query = sum_query.filter(Messages.timestamp < self.end_date)
        results = [dict(row) for row in current_app.session.execute(query).mappings()]
        sum  = current_app.session.execute(sum_query).all()[0].sum
        return {'data': results, 'columns': ['Сообщение', 'Дата'], 'customer': customer, 'sum': sum, 'order': order}

    def fill_itog(self):
        session: Session = current_app.session
        start_date = session.query(Settings).filter(Settings.name == Settings.START_DATE).one()
        start_date = dt.strptime(start_date.value, "%d.%m.%Y")
        if not session.query(Itog).filter(Itog.date == start_date, Itog.type == Itog.INIT).first():
            session.add(Itog(date=start_date, type=Itog.INIT))
            session.commit()
        # for_client = aliased(Clients)
        # query = session.query(MessageOrders).join(Messages).join(Goods) \
        #     .join(Customers).join(Clients, Customers.clients, isouter=True).join(
        #     Messages.for_client.of_type(for_client), isouter=True) \
        #     .with_entities(Goods.id.label('good_id'), func.coalesce(Goods.price, 0).label('price'),
        #                    func.coalesce(Goods.org_price, 0).label('org'),
        #                    func.coalesce(for_client.id, Clients.id).label('client_id'),
        #                    func.sum(MessageOrders.quantity).label('count')).filter(
        #     Messages.timestamp >= start_date).group_by(text("1,2,3,4")).order_by(Goods.type)
        session.query(Itog).filter(Itog.date == start_date, Itog.type == Itog.CALCULATED).delete()
        # all = query.all()
        df = self.get_itog_data(False, True)
        df = df.groupby(['client_id', 'good_id']).sum()
        for index, row in df.iterrows():
            session.add(
                Itog(date=start_date, client_id=index[0], good_id=index[1], quantity=row['count'], price=row['price'],
                     org=row['org_price'], sum=row['sum']+row['org'], type=Itog.CALCULATED))
        session.commit()
        res = session.execute(text(
            f"insert into itog (date, client_id, org, sum, type) select date, client_id, sum(org),sum(sum), {Itog.CALCULATED} from itog where date(date)=date('{start_date.strftime('%Y-%m-%d')}') and client_id is not null and good_id is not null and type={Itog.CALCULATED} group by 1,2"))
        session.execute(text(
            f"insert into itog (date, good_id,price, org, quantity, sum, type) select date, good_id, price, org,sum(quantity),sum(sum), {Itog.CALCULATED} from itog where date(date)=date('{start_date.strftime('%Y-%m-%d')}') and client_id is not null and good_id is not null and type={Itog.CALCULATED} group by 1,2,3,4"))
        session.commit()



    def get_itog_data(self, to_excel=False, for_itog=False):
        '''
        Получает сборный DataFrame с заказми и оплатами
        Args:
            to_excel: для выгрузки в файл
            for_itog: для заполнения таблицы itog, если нет то добавляются доп поля (суммы)

        Returns: DataFrame
        '''
        session = current_app.session
        if self.start_date is None:
            start_date = session.query(Settings).filter(Settings.name == Settings.START_DATE).one()
            start_date = dt.strptime(start_date.value, "%d.%m.%Y")
        else:
            start_date = self.start_date
        payments_query = select(func.coalesce(Clients.duplicate_for, Clients.id).label('client_id'),
                                func.sum(Payments.sum).label('p_sum')).join(Payers).join(payers_to_clients, isouter=True). \
            join(Clients, func.coalesce(Payments.for_client_id, payers_to_clients.c.client_id) == Clients.id) \
            .filter(Payments.timestamp >= start_date, func.coalesce(Payments.not_use, False) == False).group_by(
            text("1"))
        if self.end_date:
            payments_query = payments_query.filter(Payments.timestamp < self.end_date)

        case_qty = case(
            ((MessageOrders.quantity / 0.5 % 2) == 0,MessageOrders.quantity),
            ((MessageOrders.quantity / 0.5 % 2) != 0, MessageOrders.quantity-0.5),
        else_=0)
        #есть ли половинка
        case_half_qty = case(
            ((MessageOrders.quantity / 0.5 % 2) != 0, 1),
        else_=0)

        query = select(func.coalesce(Goods.short_name, Goods.name).label('good'),
                       Goods.id.label('good_id'),
                       func.coalesce(Goods.weight, 0).label('weight'),
                       func.cast(func.coalesce(Goods.org_price, 0), Integer).label('org_price'),
                       func.coalesce(Clients.duplicate_for, Clients.id).label('client_id'),
                       func.coalesce(Goods.active, False).label('active'),
                       # кастим во float, иначе в pandas тип будет object и в excel выгрузятся с ,
                       func.cast(func.sum(MessageOrders.quantity), Float).label('count'),
                       func.aggregate_strings(Messages.props['comment'].as_string(), ";").label('comment'),
                       func.min(func.unixepoch(Messages.timestamp)).label('timestamp'),
                       # кол-во целое * орг + половинка(если есть) * 100
                       func.sum(case((Clients.id != 139 ,case_qty * Goods.org_price + case_half_qty * 100), else_=0)).label('org')) \
                .select_from(MessageOrders) \
            .join(Messages).join(Customers).join(customers_to_clients, isouter=True).join(Goods) \
            .join(Clients, func.coalesce(Messages.for_client_id, customers_to_clients.c.client_id) == Clients.id) \
            .filter(Messages.timestamp >= start_date).group_by(text("1,2,3,4,5,6,12")).order_by(Goods.type)
        if self.end_date:
            query = query.join(Prices, and_(Prices.good_id == Goods.id, Prices.date >= start_date, Prices.date <= self.end_date), isouter=True)
            query = query.filter(Messages.timestamp < self.end_date)
            query = query.add_columns(func.sum(MessageOrders.quantity * func.coalesce(Prices.price, Goods.price,0)).label('sum'),
                                      func.cast(func.coalesce(Prices.price, Goods.price, 0), Integer).label('price'),
                                      )
        else:
            query = query.add_columns(func.sum(MessageOrders.quantity * Goods.price).label('sum'),
                                      func.cast(func.coalesce(Goods.price, 0), Integer).label('price'),
                                      )
        df = pd.DataFrame.from_records(session.execute(query).mappings())
        payments = pd.read_sql(payments_query, session.bind, index_col="client_id")
        if for_itog:
            return df
        # payments.columns = pd.MultiIndex.from_product([payments.columns, [0], [0]])
        df['weight'] = df['weight'].astype(float)
        payments.columns = pd.MultiIndex.from_product([[''], ['Оплаты'], [''], [''], ['']])
        # sum=df.pivot(index='client_id', columns=['good', 'weight', 'price', 'org_price'], values='sum')
        sum = df.groupby('client_id').sum()[['sum', 'org']]
        sum['sum'] = sum['sum'].astype('float')
        sum['org'] = sum['org'].astype('float')
        sum.columns = pd.MultiIndex.from_product([[''], sum.columns, [''], [''], ['']])
        # time_st = df.groupby('client_id').min()['timestamp'].to_frame()
        # time_st.columns = pd.MultiIndex.from_product([[''], time_st.columns, [''], [''], ['']])
        comm = df.groupby('client_id')['comment'].sum().to_frame()
        comm.columns = pd.MultiIndex.from_product([[''], comm.columns, [''], [''], ['']])
        clients = pd.DataFrame.from_records(session.execute(select(Clients.id, Clients.name)).mappings(), index='id')
        clients.columns = pd.MultiIndex.from_product([[''], clients.columns, [''], [''], ['']])
        columns = ['good_id', 'good', 'weight', 'price', 'org_price']
        df = df.pivot(index=['client_id'], columns=columns, values='count')
        if not to_excel:
            df = df.merge(sum, how="left", left_index=True, right_index=True)
        df = df.merge(comm, how="left", left_index=True, right_index=True)
        df = df.merge(payments, how="left", left_index=True, right_index=True)
        df = df.merge(clients, left_index=True, right_index=True)
        # df = df.merge(time_st, left_index=True, right_index=True)
        if not to_excel:
            df['', 'diff', '', '', ''] = df['', 'org', '', '', ''] + df['', 'sum', '', '', ''] - df['', 'Оплаты', '', '', '']
        df['', 'client_id', '', '', ''] = df.index
        df.index = df['']['name']  # ['']['']['']['']
        df = df.drop(('', 'name', '', '', ''), axis=1)
        df = df.sort_index()
        # df = df.sort_values(by=('', 'timestamp', '', '', ''))


        return df


    def create_report(self, to_excel=False):
        df = self.get_itog_data(to_excel, False)

        if to_excel:
            df.columns = df.columns.droplevel()  # убираем ид товара из первой строки
            df = df.drop(('client_id', '', '', ''), axis=1)
            # оплаты в конец
            column_to_move = df.pop(("Оплаты", '', '', ''))
            comment_col = df.pop(("comment", '', '', ''))
            # разные варианты добавления формул (сразу в фрейм или потом в excel)
            col_l = get_column_letter(len(df.columns) + 1)
            df['Стоим.', '', '', ''] = [f"=SUMPRODUCT(B$4:{col_l}$4,B{i + 6}:{col_l}{i + 6})" for i in np.arange(len(df))]
            # так работает R1C1 INDIRECT
            # df['Стоим.', '', '', ''] = f'=SUMPRODUCT(INDIRECT("R4C2",0):INDIRECT("R4C[-1]",0), INDIRECT("RC2",0):INDIRECT("RC[-1]",0))'
            df['Орг.', '', '', ''] = None
            df['К оплате', '', '', ''] = None
            df.insert(len(df.columns), "Оплаты", column_to_move)
            df.insert(len(df.columns), "comment", comment_col)

            path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'report.xlsx')
            styled = df.style.apply_index(lambda x: [f'text-align: left; font-weight: bold' for v in x], axis="rows")
            sheet_name = "Список"
            with pd.ExcelWriter(path) as wb:
                styled.to_excel(wb, sheet_name=sheet_name, startrow=1)
                sheet: Worksheet = wb.book.worksheets[0]
                sheet['A2'].value = "Имя"
                sheet['A3'].value = "Вес"
                sheet['A4'].value = "Цена"
                sheet['A5'].value = "Орг."
                sheet.delete_rows(6)
                r = sheet.max_row
                c = sheet.max_column
                # имя, стоим, орг, к оплате,  оплаты, коммент
                col_sum_l = get_column_letter(c - 4)
                col_org_l = get_column_letter(c - 3)
                col_itog_l = get_column_letter(c - 2)
                cmaxl = get_column_letter(c - 5)
                # сумма
                # for i in range(6, r + 1):
                #     sheet[f'{col_sum_l}{i}'].value = f"=SUMPRODUCT(B4:{cmaxl}4,B{i}:{cmaxl}{i})"
                # орг
                for i in range(7, r + 1):
                    # sheet[f'{col_org_l}{i}'].value = f"=SUMPRODUCT(B5:{cmaxl}5,B{i}:{cmaxl}{i})"
                    sheet[f'{col_org_l}{i}'].value = ArrayFormula(f'{col_org_l}{i}', f"=sum(if(iseven(B{i}:{cmaxl}{i}/0.5),B{i}:{cmaxl}{i},B{i}:{cmaxl}{i}-0.5)*B$5:{cmaxl}$5+if(iseven(B{i}:{cmaxl}{i}/0.5),0,1)*{self.org_half})")
                # итог по клиенту
                for i in range(6, r + 1):
                    sheet[f'{col_itog_l}{i}'].value = f"=SUM({col_sum_l}{i}:{col_org_l}{i})"
                # суммы к оплате
                for i in range(2, c):
                    col_sum_l = get_column_letter(i)
                    sheet[f'{col_sum_l}{r + 1}'].value = f"=SUM({col_sum_l}6:{col_sum_l}{r})"

        return df.to_json(orient='index')

    def compare_report(self):
        session = current_app.session
        summary = gSheets(session).get_summary()
        # summary = json.loads(summary)
        report = self.create_report()
        clients = defaultdict(lambda: defaultdict(int))
        goods = defaultdict(int)
        clients_gs = defaultdict(lambda: defaultdict(int))
        for row in summary[1]:
            if row['value'] > 0:
                clients_gs[row['имя']][row['variable']] += row['value']
        for row in report:
            clients[row['client']][row['good']] += 1
            goods[row['good']] += 1
        result = {
            'goods': goods,
            'clients': clients,
            'gs_goods': summary[2],
            'gs_clients': clients_gs
        }
        return result
